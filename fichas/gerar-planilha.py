#!/usr/bin/env python3
"""Gera a planilha de cronograma do ensaio de pré-fermentos da QT."""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/duvidas-bestas/fichas/QT-cronograma-ensaio.xlsx"

DT = "ddd dd/mm hh:mm"
DUR = "[h]:mm"
HORA = "hh:mm"

INK, PAPER, SHADE = "1A1E1E", "FFFFFF", "E4E1E0"
YELLOW = "FFFFCC"
BLUE, GREEN, BLACK = "0000FF", "008000", "1A1E1E"

F = "Arial"
f_title = Font(name=F, size=14, bold=True, color=INK)
f_sub = Font(name=F, size=10, color="6E7474")
f_bar = Font(name=F, size=9, bold=True, color=PAPER)
f_head = Font(name=F, size=9, bold=True, color=INK)
f_body = Font(name=F, size=10, color=BLACK)
f_bodyb = Font(name=F, size=10, bold=True, color=BLACK)
f_in = Font(name=F, size=10, bold=True, color=BLUE)
f_link = Font(name=F, size=10, color=GREEN)
f_note = Font(name=F, size=9, color="4A5050")

fill_bar = PatternFill("solid", fgColor=INK)
fill_in = PatternFill("solid", fgColor=YELLOW)
fill_rec = PatternFill("solid", fgColor=SHADE)

hair = Side(style="thin", color="BFC4C3")
b_bottom = Border(bottom=hair)
b_box = Border(left=hair, right=hair, top=hair, bottom=hair)

# Durações por braço, em horas. Vêm do caderno QT-ensaio-pre-fermentos rev. 2.
PAR = [
    ("Mistura da biga",                 None,  None),
    ("Arranque a 20 °C",                1.00,  0.75),
    ("Câmara a 4 °C",                  19.50, 15.00),
    ("Risveglio a 21 °C",               3.00,  2.00),
    ("Mistura do rinfresco",            0.3333, 0.3333),
    ("Puntata a 22 a 24 °C",            2.9167, 0.9167),
    ("Appretto a 18 °C",                4.75,  4.75),
    ("Checagem 1, horas de câmara",     7.50,  8.00),
    ("Checagem 2, horas de câmara",     16.50, 13.00),
]
OBS = [
    "editável, dispara todo o resto",
    "antes de ir para o frio",
    "encurtada na revisão 2 para caber no horário",
    "compensa a câmara mais curta",
    "água a fio, sal na última adição",
    "absorve a diferença entre os dois braços",
    "igual nos dois braços, de propósito",
    "leitura no testemunho de 200 g",
    "leitura no testemunho de 200 g",
]

ETAPAS = ["Mistura da biga", "Entra na câmara a 4 °C", "Checagem 1", "Checagem 2",
          "Sai da câmara", "Rinfresco, início", "Rinfresco, fim da mistura",
          "Staglio, 6 × 290 g", "Forno"]

wb = Workbook()
ws = wb.active
ws.title = "Cronograma"


def bar(sheet, row, text, last_col=4):
    sheet.cell(row=row, column=1, value=text)
    for c in range(1, last_col + 1):
        cell = sheet.cell(row=row, column=c)
        cell.fill = fill_bar
        cell.font = f_bar
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 17


ws["A1"] = "QT PIZZA BAR · ENSAIO DE PRÉ-FERMENTOS"
ws["A1"].font = f_title
ws["A2"] = "Cronograma automático · braço B (biga a 55%) e braço C (biga a 65%)"
ws["A2"].font = f_sub

ws["A4"] = "COMO USAR"
ws["A4"].font = f_head
for i, txt in enumerate([
    "Edite só as células amarelas. Mude a hora da mistura da biga e todo o cronograma daquele braço se recalcula.",
    "Células cinza são para anotar à mão durante a produção: hora real e pH medido.",
    "Durações no formato h:mm, por exemplo 19:30 para dezenove horas e meia. Mistura no formato dia/mês hora:minuto.",
], start=5):
    ws.cell(row=i, column=1, value=txt).font = f_note

# ---------------- parâmetros ----------------
bar(ws, 9, "  PARÂMETROS · o que você edita")
r = 10
for j, h in enumerate(["Parâmetro", "B · Morbida 55%", "C · Umida 65%", "Observação"], start=1):
    c = ws.cell(row=r, column=j, value=h)
    c.font, c.border = f_head, b_bottom

for k, (label, vb, vc) in enumerate(PAR):
    r = 11 + k
    ws.cell(row=r, column=1, value=label).font = f_body
    for col, val in ((2, vb), (3, vc)):
        cell = ws.cell(row=r, column=col)
        if val is None:
            cell.value = datetime(2026, 8, 5, 14, 30) if col == 2 else datetime(2026, 8, 5, 22, 15)
            cell.number_format = DT
        else:
            cell.value = val / 24.0
            cell.number_format = DUR
        cell.font, cell.fill, cell.border = f_in, fill_in, b_box
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=OBS[k]).font = f_note

r = 20
ws.cell(row=r, column=1, value="Total, da mistura ao forno").font = f_bodyb
for col in (2, 3):
    L = get_column_letter(col)
    cell = ws.cell(row=r, column=col, value=f"=SUM({L}12:{L}17)")
    cell.number_format, cell.font = DUR, f_bodyb
    cell.alignment = Alignment(horizontal="center")
ws.cell(row=r, column=4, value="soma das durações, sem contar as checagens").font = f_note


def cronograma(sheet, first_row, titulo, par_col):
    """Escreve um bloco de cronograma encadeado a partir da coluna de parâmetros."""
    bar(sheet, first_row, titulo)
    hr = first_row + 1
    for j, h in enumerate(["Etapa", "Previsto", "Hora real", "pH observado"], start=1):
        c = sheet.cell(row=hr, column=j, value=h)
        c.font, c.border = f_head, b_bottom
    P = par_col           # coluna dos parâmetros: B ou C
    r0 = hr + 1           # primeira linha de etapa
    # cada etapa se encadeia na anterior, com guarda para mistura em branco
    formulas = [
        f'=IF(${P}$11="","",${P}$11)',                       # mistura
        f'=IF($B{r0}="","",$B{r0}+${P}$12)',                 # entra na câmara
        f'=IF($B{r0+1}="","",$B{r0+1}+${P}$18)',             # checagem 1
        f'=IF($B{r0+1}="","",$B{r0+1}+${P}$19)',             # checagem 2
        f'=IF($B{r0+1}="","",$B{r0+1}+${P}$13)',             # sai da câmara
        f'=IF($B{r0+4}="","",$B{r0+4}+${P}$14)',             # rinfresco início
        f'=IF($B{r0+5}="","",$B{r0+5}+${P}$15)',             # rinfresco fim
        f'=IF($B{r0+6}="","",$B{r0+6}+${P}$16)',             # staglio
        f'=IF($B{r0+7}="","",$B{r0+7}+${P}$17)',             # forno
    ]
    for k, (etapa, fx) in enumerate(zip(ETAPAS, formulas)):
        r = r0 + k
        destaque = etapa in ("Sai da câmara", "Rinfresco, início", "Forno")
        sheet.cell(row=r, column=1, value=etapa).font = f_bodyb if destaque else f_body
        cell = sheet.cell(row=r, column=2, value=fx)
        cell.number_format = DT
        cell.font = f_bodyb if destaque else f_body
        cell.alignment = Alignment(horizontal="center")
        cell.border = b_box
        for col, fmt in ((3, HORA), (4, "0.00")):
            rec = sheet.cell(row=r, column=col)
            rec.fill, rec.border, rec.number_format, rec.font = fill_rec, b_box, fmt, f_body
            rec.alignment = Alignment(horizontal="center")
    return r0 + 8         # linha do forno


forno_b = cronograma(ws, 22, "  CRONOGRAMA · BRAÇO B · BIGA A 55% · FERMENTO 0,9%", "B")
forno_c = cronograma(ws, 34, "  CRONOGRAMA · BRAÇO C · BIGA A 65% · FERMENTO 0,55%", "C")

# ---------------- convergência ----------------
bar(ws, 46, "  CONVERGÊNCIA · as duas precisam assar no mesmo horário")
conv = [
    ("Forno do braço B",  f"=$B${forno_b}", DT, f_body),
    ("Forno do braço C",  f"=$B${forno_c}", DT, f_body),
    ("Diferença",
     f'=IF(OR($B${forno_b}="",$B${forno_c}=""),"",($B${forno_c}-$B${forno_b})*1440)',
     '0" min"', f_bodyb),
    ("Situação",
     '=IF($B48="","",IF(ABS($B48)<=10,"Convergem. Ok para assar juntas.",'
     '"Fora de sincronia. Use uma das duas linhas abaixo."))', "General", f_bodyb),
    ("Para C assar junto com B, bata a biga C às",
     f'=IF($B${forno_b}="","",$B${forno_b}-$C$20)', DT, f_bodyb),
    ("Para B assar junto com C, bata a biga B às",
     f'=IF($B${forno_c}="","",$B${forno_c}-$B$20)', DT, f_bodyb),
]
for k, (label, fx, fmt, fnt) in enumerate(conv):
    r = 47 + k
    ws.cell(row=r, column=1, value=label).font = fnt
    cell = ws.cell(row=r, column=2, value=fx)
    cell.number_format, cell.font = fmt, fnt
    cell.alignment = Alignment(horizontal="center")
    if k >= 4:
        cell.border = b_box
ws.cell(row=51, column=4,
        value="copie o valor para a célula B11 ou C11 e o cronograma se realinha").font = f_note

# ---------------- exemplo ----------------
ws["A54"] = "EXEMPLO DE PREENCHIMENTO DAS COLUNAS CINZA"
ws["A54"].font = f_head
ws["A55"] = "Sai da câmara"
ws["A55"].font = f_note
for col, val, fmt in ((2, "previsto qui 06/08 11:00", "General"), (3, 0.4736, HORA), (4, 5.08, "0.00")):
    c = ws.cell(row=55, column=col, value=val)
    c.number_format, c.font = fmt, f_note
    c.alignment = Alignment(horizontal="center")
ws["A56"] = "Anote só o relógio na coluna Hora real, sem a data. pH com duas casas."
ws["A56"].font = f_note

# ---------------- notas ----------------
bar(ws, 58, "  NOTAS QUE NÃO PODEM SUMIR NA PASSAGEM PARA PLANILHA")
notas = [
    "Appretto igual nos dois braços é decisão de desenho. Se as bolas não estiverem prontas, atrase o forno, nunca encurte o appretto.",
    "pH sempre em suspensão de 10 g de biga em 100 mL de água destilada a 20 °C. Espeto direto na massa firme dá ruído, não sinal.",
    "Chegou cedo ao ponto, baixe a câmara para 2 °C. Está atrasada, suba o risveglio de 21 °C para 24 a 26 °C.",
    "O risveglio mais longo recupera o pH, não o platô. As duas perdem profundidade aromática em proporção parecida, então a comparação segue de pé.",
    "Sem o braço A, anote a biga a 50% que você já faz como referência, senão você compara 55 contra 65 sem âncora.",
    "Hidratação final de 75% nos dois braços. Fermento fresco de 0,9% em B e 0,55% em C, sobre a farinha.",
]
for k, t in enumerate(notas):
    ws.cell(row=59 + k, column=1, value=("· " + t)).font = f_note

ws.cell(row=66, column=1,
        value="Fonte das durações e das doses: caderno QT-ensaio-pre-fermentos, revisão 2. "
              "Horários de mistura e de forno definidos por Matheus.").font = f_note

for col, w in ((1, 44), (2, 20), (3, 13), (4, 62)):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A10"

# ================= aba de trás para frente =================
ws2 = wb.create_sheet("De trás para frente")
ws2["A1"] = "A QUE HORA BATER CADA BIGA"
ws2["A1"].font = f_title
ws2["A2"] = "Você informa a hora do forno e a planilha devolve os horários. Usa as mesmas durações da aba Cronograma."
ws2["A2"].font = f_sub

ws2["A4"] = "Forno desejado"
ws2["A4"].font = f_bodyb
c = ws2["B4"]
c.value = datetime(2026, 8, 6, 22, 0)
c.number_format, c.font, c.fill, c.border = DT, f_in, fill_in, b_box
c.alignment = Alignment(horizontal="center")
ws2["C4"] = "editável"
ws2["C4"].font = f_note

bar(ws2, 6, "  HORÁRIOS CALCULADOS DE TRÁS PARA FRENTE", 3)
for j, h in enumerate(["Etapa", "B · Morbida 55%", "C · Umida 65%"], start=1):
    cc = ws2.cell(row=7, column=j, value=h)
    cc.font, cc.border = f_head, b_bottom

r0 = 8
for k, etapa in enumerate(ETAPAS):
    r = r0 + k
    destaque = etapa in ("Mistura da biga", "Forno")
    ws2.cell(row=r, column=1, value=etapa).font = f_bodyb if destaque else f_body
    for col, P in ((2, "B"), (3, "C")):
        L = get_column_letter(col)
        if k == 0:
            fx = f'=IF($B$4="","",$B$4-Cronograma!${P}$20)'
        elif k == 1:
            fx = f'=IF({L}{r0}="","",{L}{r0}+Cronograma!${P}$12)'
        elif k == 2:
            fx = f'=IF({L}${r0+1}="","",{L}${r0+1}+Cronograma!${P}$18)'
        elif k == 3:
            fx = f'=IF({L}${r0+1}="","",{L}${r0+1}+Cronograma!${P}$19)'
        elif k == 4:
            fx = f'=IF({L}${r0+1}="","",{L}${r0+1}+Cronograma!${P}$13)'
        else:
            prev = r0 + k - 1
            par = {5: 14, 6: 15, 7: 16, 8: 17}[k]
            fx = f'=IF({L}${prev}="","",{L}${prev}+Cronograma!${P}${par})'
        cell = ws2.cell(row=r, column=col, value=fx)
        cell.number_format = DT
        cell.font = f_bodyb if destaque else f_link
        cell.alignment = Alignment(horizontal="center")
        cell.border = b_box

ws2.cell(row=18, column=1,
         value="A última linha tem que bater exatamente com o forno desejado. "
               "Se não bater, alguma duração da aba Cronograma foi editada no meio do caminho.").font = f_note
ws2.cell(row=19, column=1,
         value="Verde indica valor que vem da aba Cronograma. Azul é o que você digita.").font = f_note

for col, w in ((1, 32), (2, 22), (3, 22)):
    ws2.column_dimensions[get_column_letter(col)].width = w

wb.save(OUT)
print("gerado:", OUT)
