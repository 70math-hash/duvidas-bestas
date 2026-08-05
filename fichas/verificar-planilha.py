#!/usr/bin/env python3
"""Confere a planilha de cronograma sem depender do LibreOffice.

Reimplementa em Python a cadeia que as fórmulas descrevem, resolvendo as
referências lidas do próprio arquivo, e compara o resultado com os horários
já validados na folha 3 do caderno. Falha com AssertionError se divergir.

Sobre o recalc.py da skill xlsx: a imagem base traz só libreoffice-core,
sem os filtros de documento, e por isso o LibreOffice inicia mas não abre
arquivo nenhum, devolvendo "source file could not be loaded" e travando o
recalc na abertura. Resolve com `apt-get update && apt-get install -y
libreoffice-calc` (libreoffice-writer para docx). Depois disso o recalc
roda normal. Este verificador continua valendo: ele confere se as fórmulas
dizem a coisa certa, coisa que o recalc não faz, já que fórmula errada
avalia sem erro nenhum.
"""
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook

XLSX = "/home/user/duvidas-bestas/fichas/QT-cronograma-ensaio.xlsx"

# Horários validados na folha 3 do caderno, revisão 2.
ESPERADO = {
    "B": ["05/08 14:30", "05/08 15:30", "05/08 23:00", "06/08 08:00",
          "06/08 11:00", "06/08 14:00", "06/08 14:20", "06/08 17:15", "06/08 22:00"],
    "C": ["05/08 22:15", "05/08 23:00", "06/08 07:00", "06/08 12:00",
          "06/08 14:00", "06/08 16:00", "06/08 16:20", "06/08 17:15", "06/08 22:00"],
}
BLOCO = {"B": 24, "C": 36}     # primeira linha de etapa de cada bloco
PARCOL = {"B": "B", "C": "C"}

wb = load_workbook(XLSX)        # fórmulas, sem valores
ws = wb["Cronograma"]

ref = re.compile(r"\$?([A-Z]+)\$?(\d+)")


def fmt(dt):
    """Arredonda para o minuto mais proximo, como o Excel faz ao exibir hh:mm."""
    return (dt + timedelta(seconds=30)).replace(second=0, microsecond=0).strftime("%d/%m %H:%M")


def valor(coord):
    """Resolve uma célula: datetime, duração em dias, ou fórmula da cadeia."""
    cell = ws[coord]
    v = cell.value
    if isinstance(v, (datetime, timedelta)):
        return v
    if isinstance(v, (int, float)):
        return timedelta(days=v)
    if isinstance(v, str) and v.startswith("="):
        # as fórmulas da cadeia têm a forma =IF(<x>="","",<a>+<b>) ou =IF(...,<a>)
        corpo = v.split(',')[-1].rstrip(')')
        partes = [p.strip() for p in corpo.split('+')]
        total = None
        for p in partes:
            m = ref.fullmatch(p.replace('$', '').strip())
            if not m:
                raise AssertionError(f"{coord}: nao entendi o termo {p!r}")
            alvo = valor(f"{m.group(1)}{m.group(2)}")
            total = alvo if total is None else total + alvo
        return total
    raise AssertionError(f"{coord}: conteudo inesperado {v!r}")


falhas = []
for arm, r0 in BLOCO.items():
    for k, esperado in enumerate(ESPERADO[arm]):
        coord = f"B{r0 + k}"
        obtido = fmt(valor(coord))
        marca = "ok " if obtido == esperado else "ERRO"
        if obtido != esperado:
            falhas.append(f"{arm} {coord}: esperado {esperado}, obtido {obtido}")
        print(f"  {marca} {arm} {coord}  {obtido}")

# Totais: mistura + total tem que cair exatamente no forno.
for arm, r0 in BLOCO.items():
    P = PARCOL[arm]
    total = sum((ws[f"{P}{r}"].value for r in range(12, 18)), timedelta())
    mistura = ws[f"{P}11"].value
    forno = mistura + total
    esperado = valor(f"B{r0 + 8}")
    horas = total.total_seconds() / 3600
    print(f"  {'ok ' if fmt(forno) == fmt(esperado) else 'ERRO'} {arm} total {horas:.2f} h -> {forno:%d/%m %H:%M}")
    if fmt(forno) != fmt(esperado):
        falhas.append(f"{arm}: total nao fecha com o forno")

# A aba de tras para frente tem que devolver a mistura original.
ws2 = wb["De trás para frente"]
forno_alvo = ws2["B4"].value
for arm, col in (("B", "B"), ("C", "C")):
    P = PARCOL[arm]
    total = sum((ws[f"{P}{r}"].value for r in range(12, 18)), timedelta())
    mistura = forno_alvo - total
    orig = ws[f"{P}11"].value
    print(f"  {'ok ' if fmt(mistura) == fmt(orig) else 'ERRO'} {arm} de tras para frente -> {mistura:%d/%m %H:%M}")
    if fmt(mistura) != fmt(orig):
        falhas.append(f"{arm}: volta de tras para frente nao bate com a mistura")

# Nenhuma funcao pos-2007 que o Excel exija prefixo.
proibidas = ("XLOOKUP", "XMATCH", "TEXTJOIN", "CONCAT", "IFS(", "SWITCH", "FILTER", "UNIQUE", "SEQUENCE")
for sheet in wb.worksheets:
    for linha in sheet.iter_rows():
        for cell in linha:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for f in proibidas:
                    if f in cell.value.upper():
                        falhas.append(f"{sheet.title}!{cell.coordinate}: usa {f}")

assert not falhas, "FALHAS:\n" + "\n".join(falhas)
print("\nTudo confere com a folha 3 do caderno.")
