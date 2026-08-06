#!/usr/bin/env python3
"""Gera a planilha de massa da QT com biga, blend de três farinhas.

Estrutura da receita informada por Matheus: a biga carrega 37% da farinha
total (10% Marrone + 27% Viola) e o rinfresco os outros 63% (28% Viola +
35% Verde), fechando 75% de hidratação final e 2,7% de sal.

A hidratação própria da biga fica como variável: é ela que muda entre os
braços do ensaio. Toda a água que não entra na biga sobra para o rinfresco,
então mexer nela muda quão mole fica a mistura final.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/duvidas-bestas/fichas/QT-massa-biga.xlsx"

INK, PAPER, SHADE = "1A1E1E", "FFFFFF", "E4E1E0"
YELLOW, RULE = "FFFFCC", "BFC4C3"
BLUE, GREEN, BLACK = "0000FF", "008000", "1A1E1E"
RED = "9C2A2A"
F = "Arial"

f_title = Font(name=F, size=14, bold=True, color=INK)
f_sub = Font(name=F, size=10, color="6E7474")
f_bar = Font(name=F, size=9, bold=True, color=PAPER)
f_head = Font(name=F, size=9, bold=True, color=INK)
f_body = Font(name=F, size=10, color=BLACK)
f_bodyb = Font(name=F, size=10, bold=True, color=BLACK)
f_in = Font(name=F, size=10, bold=True, color=BLUE)
f_link = Font(name=F, size=10, color=GREEN)
f_note = Font(name=F, size=9, color="4A5050")

fill_bar = PatternFill("solid", fgColor=INK)
fill_in = PatternFill("solid", fgColor=YELLOW)
fill_shade = PatternFill("solid", fgColor=SHADE)

hair = Side(style="thin", color=RULE)
b_bottom = Border(bottom=hair)
b_box = Border(left=hair, right=hair, top=hair, bottom=hair)
b_top = Border(top=Side(style="medium", color=INK))

G = '#,##0.0" g"'
G0 = '#,##0" g"'
PCT = "0.0%"
PCT2 = "0.00%"

wb = Workbook()
ws = wb.active
ws.title = "Massa"


def bar(sheet, row, text, last=4):
    sheet.cell(row=row, column=1, value=text)
    for c in range(1, last + 1):
        cell = sheet.cell(row=row, column=c)
        cell.fill, cell.font = fill_bar, f_bar
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 17


def inp(sheet, row, label, value, fmt, obs=""):
    sheet.cell(row=row, column=1, value=label).font = f_body
    c = sheet.cell(row=row, column=2, value=value)
    c.number_format, c.font, c.fill, c.border = fmt, f_in, fill_in, b_box
    c.alignment = Alignment(horizontal="center")
    sheet.cell(row=row, column=3, value=obs).font = f_note


def out(sheet, row, label, formula, fmt, bold=False, obs="", col=2):
    sheet.cell(row=row, column=1, value=label).font = f_bodyb if bold else f_body
    c = sheet.cell(row=row, column=col, value=formula)
    c.number_format = fmt
    c.font = f_bodyb if bold else f_body
    c.alignment = Alignment(horizontal="center")
    if obs:
        sheet.cell(row=row, column=3, value=obs).font = f_note


ws["A1"] = "QT PIZZA BAR · MASSA COM BIGA"
ws["A1"].font = f_title
ws["A2"] = "Blend de três farinhas · biga a 37% da farinha · 75% de hidratação final"
ws["A2"].font = f_sub

ws["A4"] = "COMO USAR"
ws["A4"].font = f_head
for i, t in enumerate([
    "Edite só as células amarelas. Todo o resto é fórmula e se recalcula sozinho.",
    "A hidratação da biga é a variável do ensaio. Mudando ela, a água se desloca da biga para o rinfresco e vice-versa: a soma continua 75%.",
    "As porcentagens de farinha são sempre sobre a farinha total, do jeito que você já conta a receita.",
], start=5):
    ws.cell(row=i, column=1, value=t).font = f_note

# ---------------- rendimento ----------------
bar(ws, 9, "  RENDIMENTO · quanto você quer produzir")
inp(ws, 10, "Número de bolas", 20, "0")
inp(ws, 11, "Peso da bola", 250, G0)
inp(ws, 12, "Perda de processo", 0.02, PCT, "resto de masseira, bancada, sobra de staglio")
out(ws, 13, "Massa total a produzir", "=B10*B11*(1+B12)", G, bold=True)

# ---------------- fórmula ----------------
bar(ws, 15, "  FÓRMULA · percentuais do padeiro, sobre a farinha total")
inp(ws, 16, "Hidratação final", 0.75, PCT, "alvo da QT e dos braços do ensaio")
inp(ws, 17, "Sal", 0.027, PCT2, "entra todo no rinfresco")
inp(ws, 18, "Fermento fresco, sobre a farinha DA BIGA", 0.009, PCT2,
    "atenção: base é a farinha da biga, não a total")
inp(ws, 19, "Biga, fração da farinha total", 0.37, PCT, "sua receita: 10% Marrone + 27% Viola")
inp(ws, 20, "Hidratação da biga", 0.55, PCT, "a variável do ensaio")
out(ws, 21, "Farinha total", "=B13/(1+B16+B17+B19*B18)", G, bold=True,
    obs="derivada do rendimento e dos percentuais")
out(ws, 22, "Água total", "=B21*B16", G)
out(ws, 23, "Sal total", "=B21*B17", G)

# ---------------- blend ----------------
bar(ws, 25, "  BLEND · percentuais sobre a farinha total")
for j, h in enumerate(["Farinha", "Na biga", "No rinfresco", "Total"], start=1):
    c = ws.cell(row=26, column=j, value=h)
    c.font, c.border = f_head, b_bottom
    if j > 1:
        c.alignment = Alignment(horizontal="center")
blend = [("Marrone", 0.10, 0.0), ("Viola", 0.27, 0.28), ("Verde", 0.0, 0.35)]
for k, (nome, nb, nr) in enumerate(blend):
    r = 27 + k
    ws.cell(row=r, column=1, value=nome).font = f_body
    for col, val in ((2, nb), (3, nr)):
        c = ws.cell(row=r, column=col, value=val)
        c.number_format, c.font, c.fill, c.border = PCT, f_in, fill_in, b_box
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=4, value=f"=B{r}+C{r}")
    c.number_format, c.font = PCT, f_bodyb
    c.alignment = Alignment(horizontal="center")
r = 30
ws.cell(row=r, column=1, value="Total").font = f_bodyb
for col in (2, 3, 4):
    L = get_column_letter(col)
    c = ws.cell(row=r, column=col, value=f"=SUM({L}27:{L}29)")
    c.number_format, c.font, c.border = PCT, f_bodyb, b_top
    c.alignment = Alignment(horizontal="center")
c = ws.cell(row=31, column=1,
            value='=IF(AND(ABS($B$30-$B$19)<0.0001,ABS($D$30-1)<0.0001),'
                  '"Blend fecha: a coluna da biga bate com a fração da biga e o total dá 100%.",'
                  '"ATENÇÃO: o blend não fecha. A coluna Na biga tem que somar a fração da biga, e o Total tem que dar 100%.")')
c.font = Font(name=F, size=9, bold=True, color=BLACK)

# ---------------- pesagem ----------------
bar(ws, 33, "  PESAGEM DA BIGA")
for j, h in enumerate(["Ingrediente", "Peso", "% s/ farinha total"], start=1):
    c = ws.cell(row=34, column=j, value=h)
    c.font, c.border = f_head, b_bottom
    if j > 1:
        c.alignment = Alignment(horizontal="center")
biga_rows = [
    ("Marrone", "=$B$21*$B$27", "=$B$27"),
    ("Viola", "=$B$21*$B$28", "=$B$28"),
    ("Verde", "=$B$21*$B$29", "=$B$29"),
    ("Água gelada", "=$B$21*$B$19*$B$20", "=$B$19*$B$20"),
    ("Fermento fresco", "=$B$21*$B$19*$B$18", "=$B$19*$B$18"),
]
for k, (nome, fx, pc) in enumerate(biga_rows):
    r = 35 + k
    ws.cell(row=r, column=1, value=nome).font = f_body
    c = ws.cell(row=r, column=2, value=fx)
    c.number_format, c.font = G, f_body
    c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=3, value=pc)
    c.number_format, c.font = PCT2, f_note
    c.alignment = Alignment(horizontal="center")
out(ws, 40, "Total da biga", "=SUM(B35:B39)", G, bold=True)
ws["B40"].border = b_top
ws.cell(row=41, column=1,
        value="Mistura curta, 4 a 5 min em v1. A biga não precisa ficar lisa.").font = f_note

bar(ws, 43, "  PESAGEM DO RINFRESCO")
for j, h in enumerate(["Ingrediente", "Peso", "% s/ farinha total"], start=1):
    c = ws.cell(row=44, column=j, value=h)
    c.font, c.border = f_head, b_bottom
    if j > 1:
        c.alignment = Alignment(horizontal="center")
rinf_rows = [
    ("Biga inteira", "=$B$40", ""),
    ("Marrone", "=$B$21*$C$27", "=$C$27"),
    ("Viola", "=$B$21*$C$28", "=$C$28"),
    ("Verde", "=$B$21*$C$29", "=$C$29"),
    ("Água", "=$B$22-$B$21*$B$19*$B$20", "=$B$16-$B$19*$B$20"),
    ("Sal", "=$B$23", "=$B$17"),
]
for k, (nome, fx, pc) in enumerate(rinf_rows):
    r = 45 + k
    ws.cell(row=r, column=1, value=nome).font = f_body
    c = ws.cell(row=r, column=2, value=fx)
    c.number_format, c.font = G, f_body
    c.alignment = Alignment(horizontal="center")
    if pc:
        c = ws.cell(row=r, column=3, value=pc)
        c.number_format, c.font = PCT2, f_note
        c.alignment = Alignment(horizontal="center")
out(ws, 51, "Total da massa", "=SUM(B45:B50)", G, bold=True)
ws["B51"].border = b_top
ws.cell(row=52, column=1,
        value="Água a fio em 2 ou 3 adições. Sal dissolvido na última. TA final 24 a 26 °C.").font = f_note

# ---------------- conferência ----------------
bar(ws, 54, "  CONFERÊNCIA E DERIVADOS")
out(ws, 55, "Fecha com a massa pedida?", "=B51-B13", G, obs="tem que dar zero")
out(ws, 56, "Hidratação sobre a farinha do rinfresco",
    "=($B$22-$B$21*$B$19*$B$20)/($B$21*(1-$B$19))", PCT, bold=True,
    obs="é o que você sente na mão ao fechar a massa")
out(ws, 57, "Farinha pré-fermentada", "=B19", PCT, obs="fração que passa pela biga")
out(ws, 58, "Fermento em % da farinha total", "=B19*B18", PCT2,
    obs="para comparar com receita que dosa na farinha total")
out(ws, 59, "Peso de bola conferido", "=B51/B10/(1+B12)", G,
    obs="tem que bater com o peso pedido")
c = ws.cell(row=60, column=1,
            value='=IF($B$22-$B$21*$B$19*$B$20<0,'
                  '"ATENÇÃO: a biga sozinha já usa mais água do que os 75% totais. Baixe a hidratação da biga ou a fração dela.",'
                  'IF(($B$22-$B$21*$B$19*$B$20)/($B$21*(1-$B$19))>1,'
                  '"ATENÇÃO: a farinha do rinfresco passa de 100% de hidratação. Vai virar caldo, suba a hidratação da biga.",'
                  '"Distribuição de água viável."))')
c.font = Font(name=F, size=9, bold=True, color=BLACK)

# ---------------- notas ----------------
bar(ws, 62, "  NOTAS")
notas = [
    "A biga carrega só 37% da farinha, então a hidratação dela mexe menos no resultado final do que mexia no ensaio, onde a biga era 100% da farinha. Não espere a mesma diferença entre 45, 55 e 65 aqui: o efeito chega diluído.",
    "O que muda de verdade nesta estrutura é a linha 56. Com a biga a 45% a farinha do rinfresco recebe perto de 93% de hidratação; a 65% ela cai para 81%. É uma massa final bem diferente de fechar, com a mesma hidratação total de 75%.",
    "Todo o fermento vai na biga. A dose é sobre a farinha DA BIGA, não sobre a total: com 37% de biga, 0,9% na biga equivale a 0,33% na farinha total.",
    "O sal entra inteiro no rinfresco, nunca na biga. Sal na biga trava a protease e você perde a maturação que foi buscar.",
    "Se mudar a fração da biga na linha 19, ajuste o blend: a coluna Na biga tem que somar exatamente essa fração, e o total das duas colunas tem que dar 100%.",
]
for k, t in enumerate(notas):
    ws.cell(row=63 + k, column=1, value="· " + t).font = f_note
ws.cell(row=69, column=1,
        value="Fonte da receita: blend e percentuais informados por Matheus. "
              "Hidratações de biga e doses de fermento: caderno QT-ensaio-pre-fermentos.").font = f_note

for col, w in ((1, 46), (2, 16), (3, 46), (4, 14)):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A9"

# ================= três braços =================
w2 = wb.create_sheet("Três braços")
w2["A1"] = "OS TRÊS BRAÇOS NA MASSA DA QT"
w2["A1"].font = f_title
w2["A2"] = ("Mesma estrutura de receita, mesma hidratação final, mesmo blend. "
            "Só muda a hidratação da biga e a dose de fermento.")
w2["A2"].font = f_sub
w2["A3"] = "Rendimento, blend, sal e hidratação final vêm da aba Massa. Edite lá e os três braços acompanham."
w2["A3"].font = f_note

bar(w2, 5, "  VARIÁVEIS DE CADA BRAÇO", 4)
for j, h in enumerate(["", "A · Fredda", "B · Morbida", "C · Umida"], start=1):
    c = w2.cell(row=6, column=j, value=h)
    c.font, c.border = f_head, b_bottom
    if j > 1:
        c.alignment = Alignment(horizontal="center")
for r, label, vals, fmt in (
    (7, "Hidratação da biga", (0.45, 0.55, 0.65), PCT),
    (8, "Fermento s/ farinha da biga", (0.010, 0.009, 0.0055), PCT2),
):
    w2.cell(row=r, column=1, value=label).font = f_body
    for k, v in enumerate(vals):
        c = w2.cell(row=r, column=2 + k, value=v)
        c.number_format, c.font, c.fill, c.border = fmt, f_in, fill_in, b_box
        c.alignment = Alignment(horizontal="center")

bar(w2, 10, "  PESAGEM · gramas por braço", 4)
for j, h in enumerate(["Ingrediente", "A · 45%", "B · 55%", "C · 65%"], start=1):
    c = w2.cell(row=11, column=j, value=h)
    c.font, c.border = f_head, b_bottom
    if j > 1:
        c.alignment = Alignment(horizontal="center")

# farinha total muda por braço porque o fermento entra no divisor
FAR = "(Massa!$B$13/(1+Massa!$B$16+Massa!$B$17+Massa!$B$19*{f}))"
linhas = [
    ("biga", "Marrone", "={FAR}*Massa!$B$27"),
    ("biga", "Viola", "={FAR}*Massa!$B$28"),
    ("biga", "Verde", "={FAR}*Massa!$B$29"),
    ("biga", "Água gelada", "={FAR}*Massa!$B$19*{h}"),
    ("biga", "Fermento fresco", "={FAR}*Massa!$B$19*{f}"),
    ("tot", "Total da biga", "=SUM({c}12:{c}16)"),
    ("rinf", "Marrone", "={FAR}*Massa!$C$27"),
    ("rinf", "Viola", "={FAR}*Massa!$C$28"),
    ("rinf", "Verde", "={FAR}*Massa!$C$29"),
    ("rinf", "Água", "={FAR}*(Massa!$B$16-Massa!$B$19*{h})"),
    ("rinf", "Sal", "={FAR}*Massa!$B$17"),
    ("tot", "Total da massa", "={c}17+SUM({c}18:{c}22)"),
]
r = 12
sub_written = set()
for grupo, nome, tpl in linhas:
    if grupo == "rinf" and "rinf" not in sub_written:
        sub_written.add("rinf")
    w2.cell(row=r, column=1, value=nome).font = f_bodyb if grupo == "tot" else f_body
    for k in range(3):
        col = 2 + k
        L = get_column_letter(col)
        h = f"{L}$7"
        f_ = f"{L}$8"
        fx = tpl.format(FAR=FAR.format(f=f_), h=h, f=f_, c=L)
        c = w2.cell(row=r, column=col, value=fx)
        c.number_format = G
        c.font = f_bodyb if grupo == "tot" else f_link
        c.alignment = Alignment(horizontal="center")
        if grupo == "tot":
            c.border = b_top
    r += 1

bar(w2, 25, "  O QUE MUDA ENTRE OS BRAÇOS", 4)
for j, h in enumerate(["", "A · 45%", "B · 55%", "C · 65%"], start=1):
    c = w2.cell(row=26, column=j, value=h)
    c.font, c.border = f_head, b_bottom
    if j > 1:
        c.alignment = Alignment(horizontal="center")
comp = [
    (27, "Hidratação sobre a farinha do rinfresco",
     "=(Massa!$B$16-Massa!$B$19*{h})/(1-Massa!$B$19)", PCT, True),
    (28, "Água na biga", "={FAR}*Massa!$B$19*{h}", G, False),
    (29, "Água no rinfresco", "={FAR}*(Massa!$B$16-Massa!$B$19*{h})", G, False),
    (30, "Fermento em % da farinha total", "=Massa!$B$19*{f}", PCT2, False),
]
for row, label, tpl, fmt, bold in comp:
    w2.cell(row=row, column=1, value=label).font = f_bodyb if bold else f_body
    for k in range(3):
        L = get_column_letter(2 + k)
        fx = tpl.format(FAR=FAR.format(f=f"{L}$8"), h=f"{L}$7", f=f"{L}$8")
        c = w2.cell(row=row, column=2 + k, value=fx)
        c.number_format = fmt
        c.font = f_bodyb if bold else f_body
        c.alignment = Alignment(horizontal="center")

w2.cell(row=32, column=1, value=(
    "A linha 27 é a leitura mais importante desta aba. A hidratação final é 75% nos três, "
    "mas a farinha que entra seca no rinfresco recebe quase 93% no braço A e 81% no braço C. "
    "São massas bem diferentes de fechar, e boa parte do que você vai sentir vem daí, "
    "não da biga em si.")).font = f_note
w2.cell(row=34, column=1, value=(
    "Verde indica valor calculado a partir da aba Massa. Azul é o que você digita.")).font = f_note

for col, w in ((1, 40), (2, 18), (3, 18), (4, 18)):
    w2.column_dimensions[get_column_letter(col)].width = w
w2.freeze_panes = "A6"

wb.save(OUT)
print("gerado:", OUT)
